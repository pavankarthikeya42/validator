"""
reporter.py — Generate JSON, CSV, Excel, and HTML Reports

Produces four output files:
  - results.json    full machine-readable results
  - results.csv     flat spreadsheet
  - results.xlsx    Excel workbook with separate sheets
  - report.html     rich human-readable report with inline screenshots
"""
from __future__ import annotations

import csv
import json
import os
import time
from pathlib import Path
from typing import Optional

from jinja2 import Environment, FileSystemLoader, select_autoescape
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from comparator import ComparisonResult, FieldStatus
from state import RunState

# Status → hex background for HTML/Excel
STATUS_COLORS = {
    FieldStatus.MATCH: "#d4edda",
    FieldStatus.PARTIAL: "#ffeeba",
    FieldStatus.MISMATCH: "#f8d7da",
    FieldStatus.MISSING_IN_PDF: "#fff3cd",
    FieldStatus.MISSING_IN_UI: "#d1ecf1",
    FieldStatus.SKIPPED: "#e2e3e5",
}

EXCEL_FILLS = {
    FieldStatus.MATCH: "D4EDDA",
    FieldStatus.PARTIAL: "FFEEBA",
    FieldStatus.MISMATCH: "F8D7DA",
    FieldStatus.MISSING_IN_PDF: "FFF3CD",
    FieldStatus.MISSING_IN_UI: "D1ECF1",
    FieldStatus.SKIPPED: "E2E3E5",
}


class Reporter:
    """Writes all four report formats to the configured output directory."""

    def __init__(self, config: dict, state: RunState) -> None:
        self.cfg = config
        self.state = state
        out_cfg = config.get("output", {})
        self.out_dir = Path(out_cfg.get("directory", "./reports"))
        self.screenshots_dir = Path(out_cfg.get("screenshots_dir", "./reports/screenshots"))
        self.json_file = self.out_dir / out_cfg.get("json_report", "results.json")
        self.csv_file = self.out_dir / out_cfg.get("csv_report", "results.csv")
        self.html_file = self.out_dir / out_cfg.get("html_report", "report.html")
        self.xlsx_file = self.out_dir / out_cfg.get("excel_report", "results.xlsx")
        self.xml_file = self.out_dir / out_cfg.get("xml_report", "results.xml")

        self.out_dir.mkdir(parents=True, exist_ok=True)
        self.screenshots_dir.mkdir(parents=True, exist_ok=True)

        self._results: list[ComparisonResult] = []

    def add_result(self, result: ComparisonResult) -> None:
        self._results.append(result)

    def save_all(self) -> None:
        self._save_json()
        self._save_csv()
        self._save_excel()
        self._save_html()
        self._save_xml()

    # ------------------------------------------------------------------
    # Summary helpers
    # ------------------------------------------------------------------
    def build_summary(self) -> dict:
        total_docs = len(self._results)
        failed = sum(1 for r in self._results if r.error)
        total_fields = sum(r.total_fields for r in self._results)
        total_matches = sum(r.matches for r in self._results)
        total_partials = sum(r.partials for r in self._results)
        total_mismatches = sum(r.mismatches for r in self._results)
        accuracy = (
            round(sum(r.accuracy for r in self._results) / len(self._results), 2)
            if self._results else 0.0
        )
        total_missing = sum(r.missing_in_pdf + r.missing_in_ui for r in self._results)
        elapsed = self.state.elapsed_seconds()
        success_rate = (
            round((total_docs - failed) / total_docs * 100, 2) if total_docs else 0.0
        )
        return {
            "total_documents_processed": total_docs,
            "total_fields_validated": total_fields,
            "total_matches": total_matches,
            "total_partials": total_partials,
            "total_mismatches": total_mismatches,
            "overall_accuracy": accuracy,
            "total_missing_fields": total_missing,
            "total_failed_documents": failed,
            "processing_time_seconds": round(elapsed, 1),
            "processing_time_human": _format_duration(elapsed),
            "success_rate_percent": success_rate,
            "run_id": self.state.run_id,
        }

    # ------------------------------------------------------------------
    # JSON
    # ------------------------------------------------------------------
    def _save_json(self) -> None:
        payload = {
            "summary": self.build_summary(),
            "documents": [r.summary_dict() for r in self._results],
        }
        with open(self.json_file, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2, default=str)

    # ------------------------------------------------------------------
    # XML
    # ------------------------------------------------------------------
    def _save_xml(self) -> None:
        import xml.etree.ElementTree as ET
        from xml.dom import minidom
        
        root = ET.Element("ValidationReport")
        
        # Summary
        summary = self.build_summary()
        sum_el = ET.SubElement(root, "Summary")
        for k, v in summary.items():
            el = ET.SubElement(sum_el, k)
            el.text = str(v)
            
        # Documents
        docs_el = ET.SubElement(root, "Documents")
        for result in self._results:
            doc_el = ET.SubElement(docs_el, "Document", id=str(result.doc_id), index=str(result.doc_index))
            if result.error:
                ET.SubElement(doc_el, "Error").text = str(result.error)
            else:
                fields_el = ET.SubElement(doc_el, "Fields")
                for fr in result.fields:
                    f_el = ET.SubElement(fields_el, "Field", name=str(fr.field_path), status=str(fr.status.value))
                    ET.SubElement(f_el, "UIValue").text = str(fr.ui_value or "")
                    ET.SubElement(f_el, "PDFValue").text = str(fr.pdf_value or "")
                
                tables_el = ET.SubElement(doc_el, "Tables")
                for t in result.tables:
                    t_el = ET.SubElement(tables_el, "Table", name=str(t.table_name), status=str(t.status.value))
                    for r in t.row_results:
                        r_el = ET.SubElement(t_el, "Row", index=str(r.row_index), status=str(r.status.value))
                        ET.SubElement(r_el, "UIValue").text = " | ".join(r.ui_row)
                        ET.SubElement(r_el, "PDFValue").text = " | ".join(r.pdf_row)
        
        xmlstr = minidom.parseString(ET.tostring(root)).toprettyxml(indent="  ")
        with open(self.xml_file, "w", encoding="utf-8") as fh:
            fh.write(xmlstr)

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------
    def _save_csv(self) -> None:
        headers = [
            "doc_index", "doc_id", "field", "ui_value", "pdf_value",
            "status", "screenshot",
        ]
        with open(self.csv_file, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=headers)
            writer.writeheader()
            for result in self._results:
                for fr in result.fields:
                    writer.writerow({
                        "doc_index": result.doc_index,
                        "doc_id": result.doc_id,
                        "field": fr.field_path,
                        "ui_value": fr.ui_value or "",
                        "pdf_value": fr.pdf_value or "",
                        "status": fr.status.value,
                        "screenshot": fr.screenshot_path or "",
                    })
                for t in result.tables:
                    for r in t.row_results:
                        writer.writerow({
                            "doc_index": result.doc_index,
                            "doc_id": result.doc_id,
                            "field": f"Table [{t.table_name}] Row {r.row_index}",
                            "ui_value": " | ".join(r.ui_row),
                            "pdf_value": " | ".join(r.pdf_row),
                            "status": r.status.value,
                            "screenshot": "",
                        })

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------
    def _save_excel(self) -> None:
        wb = openpyxl.Workbook()

        # --- Summary sheet ---
        ws_sum = wb.active
        ws_sum.title = "Summary"
        summary = self.build_summary()
        ws_sum.append(["Metric", "Value"])
        ws_sum["A1"].font = Font(bold=True)
        ws_sum["B1"].font = Font(bold=True)
        for k, v in summary.items():
            ws_sum.append([k.replace("_", " ").title(), v])
        ws_sum.column_dimensions["A"].width = 35
        ws_sum.column_dimensions["B"].width = 20

        # --- Details sheet ---
        ws_det = wb.create_sheet("Field Results")
        headers = ["Doc Index", "Doc ID", "Field", "UI Value", "PDF Value", "Status"]
        ws_det.append(headers)
        for cell in ws_det[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        for result in self._results:
            for fr in result.fields:
                row = [
                    result.doc_index,
                    result.doc_id,
                    fr.field_path,
                    fr.ui_value or "",
                    fr.pdf_value or "",
                    fr.status.value,
                ]
                ws_det.append(row)
                color = EXCEL_FILLS.get(fr.status, "FFFFFF")
                for cell in ws_det[ws_det.max_row]:
                    cell.fill = PatternFill("solid", fgColor=color)
            
            for t in result.tables:
                for r in t.row_results:
                    row = [
                        result.doc_index,
                        result.doc_id,
                        f"Table [{t.table_name}] Row {r.row_index}",
                        " | ".join(r.ui_row),
                        " | ".join(r.pdf_row),
                        r.status.value,
                    ]
                    ws_det.append(row)
                    color = EXCEL_FILLS.get(r.status, "FFFFFF")
                    for cell in ws_det[ws_det.max_row]:
                        cell.fill = PatternFill("solid", fgColor=color)

        for col in ["A", "B", "C", "D", "E", "F"]:
            ws_det.column_dimensions[col].width = 30

        # --- Mismatches sheet ---
        ws_mis = wb.create_sheet("Mismatches")
        ws_mis.append(headers)
        for cell in ws_mis[1]:
            cell.font = Font(bold=True)
        for result in self._results:
            for fr in result.fields:
                if fr.status in (
                    FieldStatus.PARTIAL,
                    FieldStatus.MISMATCH,
                    FieldStatus.MISSING_IN_PDF,
                    FieldStatus.MISSING_IN_UI,
                ):
                    ws_mis.append([
                        result.doc_index, result.doc_id, fr.field_path,
                        fr.ui_value or "", fr.pdf_value or "", fr.status.value,
                    ])
                    color = EXCEL_FILLS.get(fr.status, "FFFFFF")
                    for cell in ws_mis[ws_mis.max_row]:
                        cell.fill = PatternFill("solid", fgColor=color)
            
            for t in result.tables:
                for r in t.row_results:
                    if r.status in (
                        FieldStatus.MISMATCH,
                        FieldStatus.MISSING_IN_PDF,
                        FieldStatus.MISSING_IN_UI,
                    ):
                        ws_mis.append([
                            result.doc_index, result.doc_id, f"Table [{t.table_name}] Row {r.row_index}",
                            " | ".join(r.ui_row), " | ".join(r.pdf_row), r.status.value,
                        ])
                        color = EXCEL_FILLS.get(r.status, "FFFFFF")
                        for cell in ws_mis[ws_mis.max_row]:
                            cell.fill = PatternFill("solid", fgColor=color)

        wb.save(self.xlsx_file)

    # ------------------------------------------------------------------
    # HTML
    # ------------------------------------------------------------------
    def _save_html(self) -> None:
        template_dir = Path(__file__).parent / "templates"
        env = Environment(
            loader=FileSystemLoader(str(template_dir)),
            autoescape=select_autoescape(["html"]),
        )
        try:
            template = env.get_template("report.html.j2")
        except Exception:
            # Fallback: generate minimal HTML inline
            self._save_html_inline()
            return

        summary = self.build_summary()
        html = template.render(
            summary=summary,
            results=self._results,
            status_colors=STATUS_COLORS,
            field_status=FieldStatus,
            screenshots_dir=str(self.screenshots_dir.resolve()),
        )
        with open(self.html_file, "w", encoding="utf-8") as fh:
            fh.write(html)

    def _save_html_inline(self) -> None:
        """Fallback HTML writer (no Jinja2 template needed)."""
        summary = self.build_summary()
        lines = [
            "<!DOCTYPE html><html lang='en'><head><meta charset='UTF-8'>",
            "<title>Validation Report</title>",
            "<style>",
            "body{font-family:Inter,sans-serif;margin:0;background:#0f172a;color:#e2e8f0}",
            ".header{background:linear-gradient(135deg,#1e3a5f,#0f172a);padding:40px;text-align:center}",
            ".header h1{margin:0;font-size:2.5rem;color:#60a5fa}",
            ".header p{color:#94a3b8;margin:8px 0 0}",
            ".summary{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px;padding:32px}",
            ".card{background:#1e293b;border-radius:12px;padding:24px;text-align:center;border:1px solid #334155}",
            ".card .val{font-size:2rem;font-weight:700;color:#60a5fa}",
            ".card .lbl{font-size:0.8rem;color:#94a3b8;margin-top:4px;text-transform:uppercase;letter-spacing:.05em}",
            ".docs{padding:0 32px 32px}",
            "details{background:#1e293b;border-radius:8px;margin-bottom:8px;border:1px solid #334155}",
            "summary{padding:16px;cursor:pointer;font-weight:600;list-style:none}",
            "summary::-webkit-details-marker{display:none}",
            "table{width:100%;border-collapse:collapse;font-size:0.85rem}",
            "th{background:#334155;padding:10px;text-align:left}",
            "td{padding:8px 10px;border-bottom:1px solid #1e293b}",
            ".MATCH{background:#14532d22}.MISMATCH{background:#7f1d1d33}",
            ".MISSING_IN_PDF{background:#78350f33}.MISSING_IN_UI{background:#0c4a6e33}",
            ".badge{padding:2px 8px;border-radius:9999px;font-size:0.75rem;font-weight:600}",
            ".b-MATCH{background:#166534;color:#bbf7d0}.b-MISMATCH{background:#991b1b;color:#fecaca}",
            ".b-MISSING_IN_PDF{background:#92400e;color:#fde68a}.b-MISSING_IN_UI{background:#075985;color:#bae6fd}",
            ".b-SKIPPED{background:#374151;color:#9ca3af}",
            "</style></head><body>",
            "<div class='header'><h1>📋 Document Validation Report</h1>",
            f"<p>Run ID: {summary['run_id']} &nbsp;|&nbsp; Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}</p></div>",
            "<div class='summary'>",
        ]

        cards = [
            ("Total Documents", summary["total_documents_processed"]),
            ("Fields Validated", summary["total_fields_validated"]),
            ("✅ Matches", summary["total_matches"]),
            ("❌ Mismatches", summary["total_mismatches"]),
            ("⚠️ Missing Fields", summary["total_missing_fields"]),
            ("💥 Failed Docs", summary["total_failed_documents"]),
            ("⏱ Duration", summary["processing_time_human"]),
            ("🏆 Success Rate", f"{summary['success_rate_percent']}%"),
        ]
        for label, value in cards:
            lines.append(f"<div class='card'><div class='val'>{value}</div><div class='lbl'>{label}</div></div>")

        lines.append("</div><div class='docs'><h2 style='color:#60a5fa;padding:0 0 16px'>Document Results</h2>")

        for result in self._results:
            status_icon = "✅" if not result.has_mismatches and not result.error else ("💥" if result.error else "❌")
            lines.append(
                f"<details><summary>{status_icon} [{result.doc_index + 1}] "
                f"<strong>{result.doc_id}</strong> — "
                f"{result.matches}✓ {result.mismatches}✗ {result.missing_in_pdf + result.missing_in_ui}⚠</summary>"
            )
            if result.error:
                lines.append(f"<p style='color:#f87171;padding:16px'>Error: {result.error}</p>")
            else:
                lines.append("<table><thead><tr><th>Field</th><th>UI Value</th><th>PDF Value</th><th>Status</th></tr></thead><tbody>")
                for fr in result.fields:
                    s = fr.status.value
                    lines.append(
                        f"<tr class='{s}'><td>{fr.field_path}</td>"
                        f"<td>{fr.ui_value or '<em>—</em>'}</td>"
                        f"<td>{fr.pdf_value or '<em>—</em>'}</td>"
                        f"<td><span class='badge b-{s}'>{s}</span></td></tr>"
                    )
                for t in result.tables:
                    for r in t.row_results:
                        s = r.status.value
                        lines.append(
                            f"<tr class='{s}'><td>Table [{t.table_name}] Row {r.row_index}</td>"
                            f"<td>{' | '.join(r.ui_row)}</td>"
                            f"<td>{' | '.join(r.pdf_row)}</td>"
                            f"<td><span class='badge b-{s}'>{s}</span></td></tr>"
                        )
                lines.append("</tbody></table>")
            lines.append("</details>")

        lines.append("</div></body></html>")
        with open(self.html_file, "w", encoding="utf-8") as fh:
            fh.write("\n".join(lines))


def _format_duration(seconds: float) -> str:
    seconds = int(seconds)
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h}h {m}m {s}s"
    if m:
        return f"{m}m {s}s"
    return f"{s}s"
